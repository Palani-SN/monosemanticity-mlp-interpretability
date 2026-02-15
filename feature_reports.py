import os
import json
import torch
import numpy as np
import pandas as pd
from torch.utils.data import DataLoader, ConcatDataset
from mlp.mlp_definition import InterpretabilityMLP
from sae.sae_definition import SparseAutoencoder
from dataset.data_loader import load_excel_to_dataloader
from openpyxl.styles import Alignment
import plotly.graph_objects as go


def save_styled_excel(df, filename="circuit_trace_detailed.xlsx"):
    # 1. Create the Excel writer object
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Detailed Trace')

        # 2. Access the openpyxl worksheet object
        workbook = writer.book
        worksheet = workbook['Detailed Trace']

        # 3. Define the center alignment style
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 4. Iterate through all rows and columns to apply alignment
        # We start from row 1 to include headers
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = center_alignment

        # 5. Optional: Auto-adjust column width for better readability
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0]
                                        .column_letter].width = length + 2

    print(f"Clean, centered report saved to {filename}")


def prep_raw_data(mlp_path, sae_path, excel_files):
    # 1. Load Models
    mlp = InterpretabilityMLP()
    mlp.load_state_dict(torch.load(mlp_path))
    mlp.eval()

    sae = SparseAutoencoder(input_dim=512, dict_size=2048)
    sae.load_state_dict(torch.load(sae_path))
    sae.eval()

    # 2. Combine all datasets (Train, Val, Test)
    all_data = []
    for file in excel_files:
        # Re-using your existing loader logic
        loader = load_excel_to_dataloader(file, batch_size=64)
        all_data.append(loader.dataset)

    combined_loader = DataLoader(ConcatDataset(
        all_data), batch_size=64, shuffle=False)

    results = []
    total_count = 0
    print("Tracing logic flow through the entire circuit...")

    with torch.no_grad():
        for batch_x, batch_y in combined_loader:
            actual_batch = mlp(batch_x)
            acts = mlp.activations['layer2']
            _, features = sae(acts)

            for i in range(batch_x.size(0)):
                total_count += 1
                x = batch_x[i].cpu().numpy()
                y = batch_y[i].cpu().numpy()

                # 1. Correct Indexing Logic
                idx1 = int(x[4])
                idx2 = int(x[9])

                # 2. Extract specific values pointed to by the indices
                # Col 1 is x[0:4], Col 2 is x[5:9]
                val1 = int(x[idx1])
                val2 = int(x[5 + idx2])

                # 3. Scalar extraction for this specific sample
                expected = int(round(y.item()))
                # Index into the batch result before calling .item()
                actual_val = actual_batch[i]

                # # 4. Activation Analysis
                # top_neuron = torch.argmax(acts[i]).item()
                # top_feat = torch.argmax(features[i]).item()
                # feat_act = torch.max(features[i]).item()

                # --- Extraction: Top 10 Neurons & Top 10 Features ---
                # We capture both the ID and the raw activation magnitude
                top_n_vals, top_n_ids = torch.topk(acts[i], 10)
                top_f_vals, top_f_ids = torch.topk(features[i], 10)

                # Convert to CPU once to speed up the inner loop
                n_ids = top_n_ids.tolist()
                n_vls = top_n_vals.tolist()
                f_ids = top_f_ids.tolist()
                f_vls = top_f_vals.tolist()

                # --- Row-per-Rank Expansion ---
                # This creates a 'Long Form' dataset perfect for SQL-like analysis
                for rank in range(10):

                    results.append({
                        "sample_id": total_count,
                        # MUST be this name
                        "idx_combination": f"{idx1}-{idx2}",
                        # MUST be this name
                        "val_combination": f"{val1}-{val2}",

                        # Inputs
                        "inputs": x.tolist(),
                        "indices": (idx1, idx2),
                        "values": (val1, val2),

                        # Circuit Rank (1 = strongest)
                        "rank": rank + 1,

                        # Neuron Data
                        "neuron_id": n_ids[rank],
                        "neuron_act": round(n_vls[rank], 6),

                        # SAE Feature Data
                        "feature_id": f_ids[rank],
                        "feature_act": round(f_vls[rank], 6),

                        # Outputs
                        "outputs": y.tolist(),

                        # Expected vs Actual
                        "expected": expected,
                        "actual": f"{actual_val.item():.4f}",
                        "delta": f"{abs(actual_val.item() - expected):.4f}",
                        "equals": int(round(actual_val.item())) == expected,
                    })

    return pd.DataFrame(results)


def generate_logic_heatmap_data(df):
    # 1. Clean and Calculate
    df['neuron_act'] = pd.to_numeric(df['neuron_act'], errors='coerce')
    df['feature_act'] = pd.to_numeric(df['feature_act'], errors='coerce')
    df['co_act_energy'] = df['neuron_act'] * df['feature_act']

    # # Extract Val1 for Y-axis (assuming "val1-val2" format)
    # df['val1'] = df['val_combination'].apply(lambda x: str(x).split('-')[0])

    # 2. Pivot Table: We'll use the SUM of energy for the color scale
    # This reflects the "Total Logic Power" being applied to that specific problem
    logic_matrix = df.pivot_table(
        index='expected',
        columns='idx_combination',
        values='co_act_energy',
        aggfunc='sum'
    ).fillna(0)

    # # Export the DataFrame to an Excel file
    # file_name = 'logic_matrix.xlsx'
    # logic_matrix.to_excel(file_name, index=False)

    # 3. Generate Unique Top 10 Circuit Metadata
    hover_metadata = {}

    # Group by logic cell coordinates
    grouped = df.groupby(['expected', 'idx_combination'])

    for (v1, idx_c), group in grouped:
        # AGGREGATION STEP: Sum energy for identical Neuron-Feature pairs
        # This removes duplicates and ranks by cumulative strength
        unique_circuits = group.groupby(['neuron_id', 'feature_id'])[
            'co_act_energy'].sum().reset_index()
        unique_circuits = unique_circuits.sort_values(
            'co_act_energy', ascending=False).head(10)

        circuit_list = []
        for _, row in unique_circuits.iterrows():
            circuit_list.append({
                "n_id": row['neuron_id'],
                "f_id": row['feature_id'],
                "energy_sum": round(float(row['co_act_energy']), 4)
            })

        hover_metadata[f"{idx_c}_{v1}"] = circuit_list

    return logic_matrix, hover_metadata


def generate_logic_heatmap_html(matrix_df, metadata, filename="logic_circuit_map.html"):
    indices = matrix_df.columns.tolist()
    values = matrix_df.index.tolist()

    data_points = []
    for y, val_cat in enumerate(values):
        for x, idx_cat in enumerate(indices):
            total_energy = float(matrix_df.iloc[y, x])
            data_points.append([x, y, round(total_energy, 4)])

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Logic Circuit Heatmap</title>
        <script src="https://code.highcharts.com/highcharts.js"></script>
        <script src="https://code.highcharts.com/modules/heatmap.js"></script>
        <style>
            body {{ font-family: 'Inter', -apple-system, sans-serif; background: #fdfdfd; padding: 40px; }}
            .container-wrapper {{ 
                background: white; padding: 25px; border-radius: 15px;
                box-shadow: 0 12px 40px rgba(0,0,0,0.08); width: fit-content; margin: auto;
                border: 1px solid #eee;
            }}
            #container {{ width: 1000px; height: 650px; }}
        </style>
    </head>
    <body>
        <div class="container-wrapper">
            <h2 style="text-align: center; color: #1a1a1a; margin-bottom: 5px;">Input Logic Density Map</h2>
            <p style="text-align: center; color: #777; margin-bottom: 25px;">
                Color intensity = <b>Total Cumulative Energy</b> (Sum of N_act * F_act)
            </p>
            <div id="container"></div>
        </div>

        <script>
        const metadata = {json.dumps(metadata)};

        Highcharts.chart('container', {{
            chart: {{ type: 'heatmap', backgroundColor: '#ffffff' }},
            title: {{ text: null }},
            xAxis: {{
                categories: {json.dumps(indices)},
                title: {{ text: 'Index Combination', style: {{ color: '#444', fontWeight: 'bold' }} }},
                labels: {{ style: {{ color: '#666' }} }}
            }},
            yAxis: {{
                categories: {json.dumps(values)},
                title: {{ text: 'Input Value (V1)', style: {{ color: '#444', fontWeight: 'bold' }} }},
                reversed: true
            }},
            colorAxis: {{
                stops: [
                    [0, '#ffffff'],       // Pure white for zero
                    [0.05, '#f1f8ff'],    // Soft start
                    [0.2, '#bbdefb'],     // Light Blue
                    [0.5, '#42a5f5'],     // Mid Blue
                    [1, '#0d47a1']        // Deep "Ground Truth" Blue
                ],
                min: 0
            }},
            tooltip: {{
                useHTML: true,
                backgroundColor: '#ffffff',
                borderWidth: 1,
                borderColor: '#ccc',
                shadow: true,
                formatter: function () {{
                    const key = this.series.xAxis.categories[this.point.x] + '_' + this.series.yAxis.categories[this.point.y];
                    const circuits = metadata[key] || [];
                    
                    let html = '<div style="padding:8px; min-width:220px">';
                    html += '<b style="color:#333">Top Unique Circuits (Summed)</b><br/>';
                    html += '<table style="width:100%; border-collapse: collapse; margin-top:8px; font-size:12px;">';
                    html += '<tr style="border-bottom: 2px solid #eee; text-align:left"><th>Pair</th><th style="text-align:right">Total Energy</th></tr>';
                    
                    circuits.forEach(c => {{
                        html += `<tr style="border-bottom: 1px solid #f9f9f9">
                            <td style="padding:4px 0">${{c.n_id}} ↔ ${{c.f_id}}</td>
                            <td style="padding:4px 0; text-align:right"><b>${{c.energy_sum}}</b></td>
                        </tr>`;
                    }});
                    html += '</table></div>';
                    return html;
                }}
            }},
            series: [{{
                name: 'Cumulative Energy',
                data: {json.dumps(data_points)},
                borderWidth: 1,
                borderColor: '#ffffff',
                dataLabels: {{
                    enabled: true,
                    color: 'contrast',
                    style: {{ textOutline: 'none', fontSize: '11px', fontWeight: 'bold', textShadow: 'none' }},
                    formatter: function() {{ return this.point.value > 0 ? this.point.value.toFixed(1) : ''; }}
                }}
            }}]
        }});
        </script>
    </body>
    </html>
    """
    # Open with explicit utf-8 encoding to support logic symbols like ↔
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html_content)
    return os.path.abspath(filename)


def generate_sankey_diagram_gchart(df, top_k=500):
    df = df.copy()
    # 1. Calculate Energy
    df['energy'] = pd.to_numeric(
        df['neuron_act']) * pd.to_numeric(df['feature_act'])

    # 2. Strict Level Prefixing (To keep 5 columns)
    df['L1'] = "(Idx) " + df['idx_combination'].astype(str)
    df['L2'] = "(Val) " + df['val_combination'].astype(str)
    df['L3'] = "(Neu) " + df['neuron_id'].astype(str)
    df['L4'] = "(Feat) " + df['feature_id'].astype(str)
    df['L5'] = "(Out) " + df['expected'].astype(str)

    flow_steps = [('L1', 'L2'), ('L2', 'L3'), ('L3', 'L4'), ('L4', 'L5')]
    rows = []

    # SCALING FACTOR: Boost small energy values so lines appear thick
    # Adjust this multiplier (e.g., 100 or 1000) based on your data range
    scale_factor = 100

    # Increase the height dynamically based on the number of unique nodes
    # If top_k is large, we want a taller canvas.
    canvas_height = 1500  # Increased from 850 for a "Deep-Dive" view

    for src_col, tgt_col in flow_steps:
        step_data = df.groupby([src_col, tgt_col])[
            'energy'].sum().reset_index()
        step_data.columns = ['source', 'target', 'value']

        # Take the most important signals
        top_step = step_data.nlargest(top_k // 4, 'value')

        for _, row in top_step.iterrows():
            if row['value'] > 0:
                # Multiply by scale_factor to ensure visual thickness
                weighted_val = float(row['value']) * scale_factor
                rows.append([row['source'], row['target'],
                            round(weighted_val, 4)])

    # 3. HTML with Enhanced Sharpness
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <script type="text/javascript" src="https://www.gstatic.com/charts/loader.js"></script>
        <style>
            body {{ font-family: 'Inter', sans-serif; background: #ffffff; padding: 30px; }}
            #sankey_main {{ width: 100%; height: {canvas_height}px; min-width: 1400px; }}
            .header {{ text-align: center; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2 style="color: #2c3e50;">Structural Interpretability: High-Fidelity Flow</h2>
            <p style="color: #7f8c8d;">Energy values scaled x{scale_factor} for visual clarity</p>
        </div>
        <div id="sankey_main"></div>

        <script type="text/javascript">
          google.charts.load('current', {{'packages':['sankey']}});
          google.charts.setOnLoadCallback(drawChart);

          function drawChart() {{
            var data = new google.visualization.DataTable();
            data.addColumn('string', 'From');
            data.addColumn('string', 'To');
            data.addColumn('number', 'Weight');
            data.addRows({json.dumps(rows)});

            // Using a Sharper, Vibrant Palette
            var colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', 
                          '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf'];

            var options = {{
              height: {canvas_height},
              sankey: {{
                iterations: 0, // Forces strict column segregation
                node: {{
                  colors: colors,
                  label: {{ fontSize: 10, color: '#1a1a1a', bold: true }},
                  width: 30,           // Thicker nodes
                  labelPadding: 25,    // More space for labels
                  interactivity: true
                }},
                link: {{
                  colorMode: 'gradient',
                  fillOpacity: 0.5,     // BOOSTED OPACITY for "Sharp" look
                  colors: colors
                }}
              }}
            }};

            var chart = new google.visualization.Sankey(document.getElementById('sankey_main'));
            chart.draw(data, options);
          }}
        </script>
    </body>
    </html>
    """

    filename = "uhd_bold_sankey.html"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html_content)
    return os.path.abspath(filename)


def generate_sankey_diagram_plotly(df, top_k=1000):
    df = df.copy()

    # 1. Calculate Energy
    df['energy'] = pd.to_numeric(
        df['neuron_act']) * pd.to_numeric(df['feature_act'])

    # 2. Strict Level Prefixing
    df['L1'] = df['idx_combination'].astype(str) + "(Idx)"
    df['L2'] = df['val_combination'].astype(str) + "(Val)"
    df['L3'] = df['neuron_id'].astype(str) + "(Neu)"
    df['L4'] = df['feature_id'].astype(str) + "(Feat)"
    df['L5'] = df['expected'].astype(str) + "(Out)"

    # Export the DataFrame to an Excel file
    # file_name = 'sankey_raw_data.xlsx'
    sorted_df = df[['L1', 'L2', 'L3', 'L4', 'L5', 'energy']
                   ].sort_values(by=['L1', 'L2', 'L3', 'L4', 'L5'])
    # sorted_df.to_excel(file_name, index=False)

    flow_steps = [('L1', 'L2'), ('L2', 'L3'), ('L3', 'L4'), ('L4', 'L5')]

    # Collect all flows with energies
    all_flows = []
    for src_col, tgt_col in flow_steps:
        step_data = sorted_df.groupby([src_col, tgt_col])[
            'energy'].sum().reset_index()
        step_data.columns = ['source', 'target', 'value']
        top_step = step_data.nlargest(top_k // 4, 'value')

        for _, row in top_step.iterrows():
            if row['value'] > 0:
                all_flows.append({
                    'source': row['source'],
                    'target': row['target'],
                    'value': float(row['value'])
                })

    # Build node list with explicit layer assignments
    layer_map = {
        '(Idx)': 0,
        '(Val)': 1,
        '(Neu)': 2,
        '(Feat)': 3,
        '(Out)': 4
    }

    # Collect all unique nodes and assign them to layers
    all_nodes = {}
    node_list = []
    node_index = 0

    for flow in all_flows:
        for node_name in [flow['source'], flow['target']]:
            if node_name not in all_nodes:
                # Determine layer from keyword
                layer = None
                for keyword, layer_num in layer_map.items():
                    if keyword in node_name:
                        layer = layer_num
                        break

                if layer is not None:
                    all_nodes[node_name] = {
                        'index': node_index,
                        'layer': layer,
                        'label': node_name
                    }
                    node_list.append(node_name)
                    node_index += 1

    # Sort nodes by layer for cleaner display
    node_list.sort(key=lambda x: all_nodes[x]['layer'])

    # Rebuild index after sorting
    for i, node_name in enumerate(node_list):
        all_nodes[node_name]['index'] = i

    # Create links (source_idx, target_idx, value)
    source_indices = []
    target_indices = []
    values = []

    for flow in all_flows:
        if flow['source'] in all_nodes and flow['target'] in all_nodes:
            source_indices.append(all_nodes[flow['source']]['index'])
            target_indices.append(all_nodes[flow['target']]['index'])
            values.append(flow['value'])

    # Assign X positions (layer positions)
    x_positions = [all_nodes[node]['layer'] / 4 for node in node_list]

    # Assign Y positions (vertical spacing within layer)
    y_positions = []
    layer_counts = {i: 0 for i in range(5)}
    layer_nodes = {i: [] for i in range(5)}

    for node in node_list:
        layer = all_nodes[node]['layer']
        layer_nodes[layer].append(node)

    for node in node_list:
        layer = all_nodes[node]['layer']
        total_in_layer = len(layer_nodes[layer])
        position_in_layer = layer_counts[layer]

        if total_in_layer > 1:
            y_pos = position_in_layer / (total_in_layer - 1)
        else:
            y_pos = 0.25

        y_positions.append(y_pos)
        layer_counts[layer] += 1

    # Pastel color palette
    pastel_colors = [
        '#A8C5E3',  # Soft Blue for Idx
        '#FFB88C',  # Soft Orange for Val
        '#B5D4A8',  # Soft Green for Neu
        '#E3A8C5',  # Soft Pink for Feat
        '#C5E3E8'   # Soft Cyan for Out
    ]

    node_colors = [pastel_colors[all_nodes[node]['layer']]
                   for node in node_list]

    # Create link colors
    link_colors = []
    for src_idx in source_indices:
        source_layer = all_nodes[node_list[src_idx]]['layer']
        base_color = pastel_colors[source_layer]
        r = int(base_color[1:3], 16)
        g = int(base_color[3:5], 16)
        b = int(base_color[5:7], 16)
        link_colors.append(f'rgba({r},{g},{b},0.3)')

    # Calculate dynamic height
    max_nodes_per_layer = max(len(layer_nodes[i]) for i in range(5))
    plot_height = max(3000, max_nodes_per_layer * 35)

    # Prepare filter options for each layer
    layer_filter_options = {}
    for layer_num in range(5):
        layer_filter_options[layer_num] = sorted(layer_nodes[layer_num])

    # Create Plotly Sankey
    fig = go.Figure(data=[go.Sankey(
        arrangement='snap',
        node=dict(
            pad=20,
            thickness=25,
            line=dict(color="white", width=2),
            label=node_list,
            color=node_colors,
            x=x_positions,
            y=y_positions,
            hovertemplate='%{label}<br>Connections: %{value}<extra></extra>',
            customdata=node_list  # Store node names for filtering
        ),
        link=dict(
            source=source_indices,
            target=target_indices,
            value=values,
            color=link_colors,
            hovertemplate='%{source.label} → %{target.label}<br>Weight: %{value:.2f}<extra></extra>'
        )
    )])

    fig.update_layout(
        font=dict(size=11, family='Arial, sans-serif', color='#333'),
        height=plot_height,
        width=1900,
        plot_bgcolor='#F8F9FA',
        paper_bgcolor='#F8F9FA',
        margin=dict(l=50, r=50, t=80, b=50)
    )

    # Generate Plotly HTML
    plotly_html = fig.to_html(include_plotlyjs='cdn', div_id='sankey_plot')

    # Prepare data for JavaScript
    all_flows_json = json.dumps(all_flows)
    layer_filter_options_json = json.dumps(layer_filter_options)

    # HTML with interactive filters
    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Sankey Flow Diagram - Interactive</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: Arial, sans-serif;
            background: #F8F9FA;
            overflow-x: hidden;
        }}
        .header {{
            position: sticky;
            top: 0;
            background: white;
            padding: 20px 30px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.08);
            z-index: 1000;
        }}
        .header h1 {{
            color: #2c3e50;
            font-size: 22px;
            font-weight: 600;
            margin: 0 0 8px 0;
        }}
        .stats {{
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
            margin-bottom: 15px;
        }}
        .stat-item {{
            color: #6c757d;
            font-size: 13px;
            padding: 6px 12px;
            background: #f1f3f5;
            border-radius: 6px;
        }}
        .stat-value {{
            font-weight: 600;
            color: #495057;
        }}
        .filters {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 15px;
            margin-top: 15px;
            padding-top: 15px;
            border-top: 1px solid #e9ecef;
        }}
        .filter-group {{
            display: flex;
            flex-direction: column;
        }}
        .filter-label {{
            font-size: 12px;
            font-weight: 600;
            color: #495057;
            margin-bottom: 6px;
        }}
        .filter-select {{
            width: 100%;
            padding: 8px;
            border: 1px solid #ced4da;
            border-radius: 4px;
            font-size: 12px;
            background: white;
            max-height: 150px;
        }}
        .filter-controls {{
            display: flex;
            gap: 10px;
            margin-top: 10px;
        }}
        .btn {{
            padding: 8px 16px;
            border: none;
            border-radius: 4px;
            font-size: 13px;
            cursor: pointer;
            font-weight: 500;
            transition: all 0.2s;
        }}
        .btn-primary {{
            background: #007bff;
            color: white;
        }}
        .btn-primary:hover {{
            background: #0056b3;
        }}
        .btn-secondary {{
            background: #6c757d;
            color: white;
        }}
        .btn-secondary:hover {{
            background: #545b62;
        }}
        .plot-container {{
            padding: 0;
            background: #F8F9FA;
        }}
        .plot-wrapper {{
            width: 100%;
            overflow: auto;
            background: #F8F9FA;
        }}
        #sankey_plot {{
            min-height: {plot_height}px;
        }}
        .layer-labels {{
            display: flex;
            justify-content: space-around;
            padding: 15px 50px;
            background: white;
            border-top: 1px solid #e9ecef;
        }}
        .layer-label {{
            text-align: center;
            font-size: 12px;
            color: #6c757d;
        }}
        .layer-name {{
            font-weight: 600;
            color: #495057;
            display: block;
            margin-bottom: 4px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Structural Interpretability Flow - Interactive Filtering</h1>
        <div class="stats">
            <div class="stat-item">
                <span class="stat-value">{len(node_list)}</span> Nodes
            </div>
            <div class="stat-item">
                <span class="stat-value">{len(values)}</span> Connections
            </div>
            <div class="stat-item">
                Layer 1: <span class="stat-value">{layer_counts[0]}</span> Idx
            </div>
            <div class="stat-item">
                Layer 2: <span class="stat-value">{layer_counts[1]}</span> Val
            </div>
            <div class="stat-item">
                Layer 3: <span class="stat-value">{layer_counts[2]}</span> Neu
            </div>
            <div class="stat-item">
                Layer 4: <span class="stat-value">{layer_counts[3]}</span> Feat
            </div>
            <div class="stat-item">
                Layer 5: <span class="stat-value">{layer_counts[4]}</span> Out
            </div>
        </div>
        
        <div class="filters">
            <div class="filter-group">
                <label class="filter-label">Layer 1 - Index</label>
                <select id="filter-layer-0" class="filter-select" multiple>
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Layer 2 - Value</label>
                <select id="filter-layer-1" class="filter-select" multiple>
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Layer 3 - Neuron</label>
                <select id="filter-layer-2" class="filter-select" multiple>
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Layer 4 - Feature</label>
                <select id="filter-layer-3" class="filter-select" multiple>
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Layer 5 - Output</label>
                <select id="filter-layer-4" class="filter-select" multiple>
                </select>
            </div>
        </div>
        
        <div class="filter-controls">
            <button class="btn btn-primary" onclick="applyFilters()">Apply Filters</button>
            <button class="btn btn-secondary" onclick="resetFilters()">Reset All</button>
            <button class="btn btn-secondary" onclick="selectAllInLayer(0)">Select All L1</button>
            <button class="btn btn-secondary" onclick="selectAllInLayer(1)">Select All L2</button>
            <button class="btn btn-secondary" onclick="selectAllInLayer(2)">Select All L3</button>
            <button class="btn btn-secondary" onclick="selectAllInLayer(3)">Select All L4</button>
            <button class="btn btn-secondary" onclick="selectAllInLayer(4)">Select All L5</button>
        </div>
    </div>
    
    <div class="plot-container">
        <div class="plot-wrapper">
            {plotly_html}
        </div>
    </div>
    
    <div class="layer-labels">
        <div class="layer-label">
            <span class="layer-name">Layer 1</span>
            <span>Index Combinations</span>
        </div>
        <div class="layer-label">
            <span class="layer-name">Layer 2</span>
            <span>Value Combinations</span>
        </div>
        <div class="layer-label">
            <span class="layer-name">Layer 3</span>
            <span>Neuron IDs</span>
        </div>
        <div class="layer-label">
            <span class="layer-name">Layer 4</span>
            <span>Feature IDs</span>
        </div>
        <div class="layer-label">
            <span class="layer-name">Layer 5</span>
            <span>Expected Output</span>
        </div>
    </div>

    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <script>
        // Data from Python
        const allFlows = {all_flows_json};
        const layerFilterOptions = {layer_filter_options_json};
        const layerMap = {{
            '(Idx)': 0,
            '(Val)': 1,
            '(Neu)': 2,
            '(Feat)': 3,
            '(Out)': 4
        }};
        
        function getNodeLayer(nodeName) {{
            for (let keyword in layerMap) {{
                if (nodeName.includes(keyword)) {{
                    return layerMap[keyword];
                }}
            }}
            return -1;
        }}
        
        // Populate filter dropdowns
        function populateFilters() {{
            for (let layer = 0; layer < 5; layer++) {{
                const select = document.getElementById(`filter-layer-${{layer}}`);
                select.innerHTML = '';
                
                layerFilterOptions[layer].forEach(node => {{
                    const option = document.createElement('option');
                    option.value = node;
                    option.textContent = node;
                    option.selected = true; // All selected by default
                    select.appendChild(option);
                }});
            }}
        }}
        
        function selectAllInLayer(layer) {{
            const select = document.getElementById(`filter-layer-${{layer}}`);
            for (let option of select.options) {{
                option.selected = true;
            }}
        }}
        
        function resetFilters() {{
            for (let layer = 0; layer < 5; layer++) {{
                selectAllInLayer(layer);
            }}
            applyFilters();
        }}
        
        function applyFilters() {{
            // Get selected nodes for each layer
            const selectedNodes = {{}};
            for (let layer = 0; layer < 5; layer++) {{
                const select = document.getElementById(`filter-layer-${{layer}}`);
                selectedNodes[layer] = Array.from(select.selectedOptions).map(opt => opt.value);
            }}
            
            // Filter flows based on selected nodes
            const filteredFlows = allFlows.filter(flow => {{
                const sourceLayer = getNodeLayer(flow.source);
                const targetLayer = getNodeLayer(flow.target);
                
                return selectedNodes[sourceLayer].includes(flow.source) && 
                       selectedNodes[targetLayer].includes(flow.target);
            }});
            
            // Rebuild node list from filtered flows
            const nodeSet = new Set();
            filteredFlows.forEach(flow => {{
                nodeSet.add(flow.source);
                nodeSet.add(flow.target);
            }});
            
            const nodeList = Array.from(nodeSet).sort((a, b) => {{
                return getNodeLayer(a) - getNodeLayer(b);
            }});
            
            // Build node mapping
            const nodeIndices = {{}};
            nodeList.forEach((node, idx) => {{
                nodeIndices[node] = idx;
            }});
            
            // Build links
            const sources = [];
            const targets = [];
            const values = [];
            
            filteredFlows.forEach(flow => {{
                if (nodeIndices[flow.source] !== undefined && nodeIndices[flow.target] !== undefined) {{
                    sources.push(nodeIndices[flow.source]);
                    targets.push(nodeIndices[flow.target]);
                    values.push(flow.value);
                }}
            }});
            
            // Calculate positions
            const xPositions = nodeList.map(node => getNodeLayer(node) / 4);
            
            const layerNodes = {{0: [], 1: [], 2: [], 3: [], 4: []}};
            nodeList.forEach(node => {{
                const layer = getNodeLayer(node);
                layerNodes[layer].push(node);
            }});
            
            const yPositions = [];
            const layerCounts = {{0: 0, 1: 0, 2: 0, 3: 0, 4: 0}};
            
            nodeList.forEach(node => {{
                const layer = getNodeLayer(node);
                const totalInLayer = layerNodes[layer].length;
                const positionInLayer = layerCounts[layer];
                
                let yPos;
                if (totalInLayer > 1) {{
                    yPos = positionInLayer / (totalInLayer - 1);
                }} else {{
                    yPos = 0.25;
                }}
                
                yPositions.push(yPos);
                layerCounts[layer]++;
            }});
            
            // Colors
            const pastelColors = ['#A8C5E3', '#FFB88C', '#B5D4A8', '#E3A8C5', '#C5E3E8'];
            const nodeColors = nodeList.map(node => pastelColors[getNodeLayer(node)]);
            
            const linkColors = sources.map(srcIdx => {{
                const sourceLayer = getNodeLayer(nodeList[srcIdx]);
                const baseColor = pastelColors[sourceLayer];
                const r = parseInt(baseColor.substr(1, 2), 16);
                const g = parseInt(baseColor.substr(3, 2), 16);
                const b = parseInt(baseColor.substr(5, 2), 16);
                return `rgba(${{r}},${{g}},${{b}},0.3)`;
            }});
            
            // Update plot
            const update = {{
                'node.label': [nodeList],
                'node.color': [nodeColors],
                'node.x': [xPositions],
                'node.y': [yPositions],
                'link.source': [sources],
                'link.target': [targets],
                'link.value': [values],
                'link.color': [linkColors]
            }};
            
            Plotly.restyle('sankey_plot', update, 0);
            
            console.log(`Filtered to ${{nodeList.length}} nodes and ${{filteredFlows.length}} connections`);
        }}
        
        // Initialize
        populateFilters();
    </script>
</body>
</html>
"""

    filename = "sankey_flow_diagram.html"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"Sankey diagram saved to: {os.path.abspath(filename)}")
    print(f"   Total nodes: {len(node_list)}")
    print(f"   Total links: {len(values)}")
    print(f"   Plot height: {plot_height}px")
    print("   Layer distribution:")
    for layer_num in range(5):
        layer_name = ['Idx', 'Val', 'Neu', 'Feat', 'Out'][layer_num]
        print(
            f"      Layer {layer_num} ({layer_name}): {layer_counts[layer_num]} nodes")

    return os.path.abspath(filename)


def generate_stacked_norm_dist(df, top_k=1000):
    """
    Generates a Highcharts Bell Curve plot with dynamic client-side generation.
    Stores only mu, sigma, count - curves generated in browser.
    """
    # 1. Prepare Numeric Data
    df['neuron_act'] = pd.to_numeric(df['neuron_act'], errors='coerce')
    df['feature_act'] = pd.to_numeric(df['feature_act'], errors='coerce')
    df['energy'] = df['neuron_act'] * df['feature_act']

    # 2. Statistical Aggregation
    stats = df.groupby(['neuron_id', 'feature_id'])['energy'].agg(
        ['mean', 'std', 'count']).reset_index()

    # Filter for validity
    stats = stats[stats['count'] > 2].dropna()

    # 3. Handle top_k logic
    if top_k is not None:
        stats['impact'] = stats['mean'] * stats['count']
        plot_stats = stats.nlargest(top_k, 'impact')
    else:
        plot_stats = stats.copy()
    
    # SORT BY MEAN DESCENDING
    plot_stats = plot_stats.sort_values('mean', ascending=False)

    # 4. Prepare lightweight data (only mu, sigma, count)
    curve_metadata = []
    for _, row in plot_stats.iterrows():
        curve_metadata.append({
            "id": f"{row['neuron_id']}__{row['feature_id']}",
            "name": f"{row['neuron_id']} ↔ {row['feature_id']}",
            "mu": round(float(row['mean']), 4),
            "sigma": round(float(row['std']), 4),
            "count": int(row['count'])
        })

    # 5. HTML Template with dynamic generation
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Circuit Stability Distributions - Interactive</title>
        <script src="https://code.highcharts.com/highcharts.js"></script>
        <style>
            body {{ 
                font-family: 'Inter', -apple-system, sans-serif; 
                background: #fcfcfc; 
                padding: 20px; 
                margin: 0;
            }}
            .container {{ 
                background: white; 
                padding: 25px; 
                border-radius: 12px; 
                border: 1px solid #eee; 
                box-shadow: 0 4px 20px rgba(0,0,0,0.05); 
                max-width: 1400px;
                margin: 0 auto;
            }}
            .controls {{
                display: grid;
                grid-template-columns: 2fr 1fr auto;
                gap: 20px;
                margin-bottom: 25px;
                align-items: end;
            }}
            .control-group {{
                display: flex;
                flex-direction: column;
            }}
            .control-label {{
                font-size: 13px;
                font-weight: 600;
                color: #495057;
                margin-bottom: 8px;
            }}
            .control-select {{
                padding: 10px;
                border: 1px solid #ced4da;
                border-radius: 6px;
                font-size: 13px;
                background: white;
                min-height: 150px;
            }}
            .control-input {{
                padding: 10px;
                border: 1px solid #ced4da;
                border-radius: 6px;
                font-size: 13px;
                width: 100%;
            }}
            .btn {{
                padding: 10px 24px;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 600;
                cursor: pointer;
                transition: all 0.2s;
                height: fit-content;
            }}
            .btn-primary {{
                background: #007bff;
                color: white;
            }}
            .btn-primary:hover {{
                background: #0056b3;
            }}
            .btn-secondary {{
                background: #6c757d;
                color: white;
                margin-top: 10px;
            }}
            .btn-secondary:hover {{
                background: #545b62;
            }}
            .info-box {{
                background: #f8f9fa;
                padding: 12px;
                border-radius: 6px;
                font-size: 12px;
                color: #6c757d;
                margin-bottom: 20px;
            }}
            #chart-div {{ 
                width: 100%; 
                height: 750px; 
            }}
            .stat-badge {{
                display: inline-block;
                background: #e7f3ff;
                color: #0056b3;
                padding: 4px 10px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 600;
                margin-right: 10px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h2 style="text-align: center; color: #1a1a1a; margin-bottom: 5px;">
                Circuit Reliability Analysis - Interactive
            </h2>
            <p style="text-align: center; color: #777; margin-bottom: 25px;">
                Circuit Reliability = <b>μ (Mean) & σ (Std Dev)</b> from N_act × F_act distributions
            </p>
            
            <div class="info-box">
                <span class="stat-badge">Total Circuits: {len(curve_metadata)}</span>
                Select circuits from the dropdown (sorted by mean ↓) and adjust sample points (10-300) to visualize their probability distributions.
                Curves are generated dynamically using the normal distribution formula.
            </div>
            
            <div class="controls">
                <div class="control-group">
                    <label class="control-label">Select Circuits (Neuron ↔ Feature) - Sorted by Mean ↓</label>
                    <select id="circuit-select" class="control-select" multiple>
                        <!-- Populated by JavaScript -->
                    </select>
                </div>
                
                <div class="control-group">
                    <label class="control-label">Sample Points (10-300)</label>
                    <input 
                        type="number" 
                        id="sample-points" 
                        class="control-input" 
                        value="50" 
                        min="10" 
                        max="300"
                        step="10"
                    />
                    <small style="color: #6c757d; margin-top: 5px;">
                        Higher = smoother curves
                    </small>
                </div>
                
                <div class="control-group">
                    <button class="btn btn-primary" onclick="plotSelectedCurves()">
                        Plot Curves
                    </button>
                    <button class="btn btn-secondary" onclick="selectAllCircuits()">
                        Select All
                    </button>
                    <button class="btn btn-secondary" onclick="clearSelection()">
                        Clear All
                    </button>
                </div>
            </div>
            
            <div id="chart-div"></div>
        </div>

        <script>
        // Curve metadata from Python (already sorted by mean descending)
        const curveMetadata = {json.dumps(curve_metadata)};
        
        let chart = null;
        
        // Initialize
        function init() {{
            // Populate circuit selector (already sorted by mean in descending order)
            const select = document.getElementById('circuit-select');
            curveMetadata.forEach(curve => {{
                const option = document.createElement('option');
                option.value = curve.id;
                option.textContent = `${{curve.name}} (μ=${{curve.mu}}, σ=${{curve.sigma}}, n=${{curve.count}})`;
                select.appendChild(option);
            }});
            
            // Initialize empty chart
            initChart();
            
            // Plot first 10 by default
            const firstTen = curveMetadata.slice(0, Math.min(10, curveMetadata.length));
            firstTen.forEach(curve => {{
                const option = select.querySelector(`option[value="${{curve.id}}"]`);
                if (option) option.selected = true;
            }});
            plotSelectedCurves();
        }}
        
        function initChart() {{
            chart = Highcharts.chart('chart-div', {{
                chart: {{ 
                    type: 'areaspline', 
                    backgroundColor: '#ffffff',
                    zoomType: 'x',
                    resetZoomButton: {{ position: {{ align: 'right', verticalAlign: 'top' }} }}
                }},
                title: {{ text: null }},
                xAxis: {{ 
                    title: {{ text: 'Activation Energy', style: {{ fontWeight: '600' }} }},
                    gridLineWidth: 1,
                    gridLineColor: '#f0f0f0'
                }},
                yAxis: {{ 
                    title: {{ text: 'Probability Density', style: {{ fontWeight: '600' }} }},
                    gridLineColor: '#f0f0f0'
                }},
                legend: {{
                    layout: 'horizontal',
                    align: 'center',
                    verticalAlign: 'bottom',
                    itemStyle: {{ fontSize: '10px', fontWeight: '400' }},
                    maxHeight: 120
                }},
                tooltip: {{
                    shared: false,
                    followPointer: true,
                    useHTML: true,
                    backgroundColor: 'rgba(255, 255, 255, 0.95)',
                    borderRadius: 8,
                    padding: 12,
                    headerFormat: '<div style="margin-bottom:8px;"><span style="color:{{series.color}};font-size:16px;">●</span> <b>{{series.name}}</b></div>',
                    pointFormat: `
                        <table style="font-size:12px;">
                            <tr><td style="padding:2px 8px 2px 0;">Density:</td><td style="font-weight:600;">{{point.y:.5f}}</td></tr>
                            <tr><td style="padding:2px 8px 2px 0;">Energy:</td><td style="font-weight:600;">{{point.x:.4f}}</td></tr>
                            <tr><td style="padding:2px 8px 2px 0;">Mean (μ):</td><td style="font-weight:600;">{{series.options.mu}}</td></tr>
                            <tr><td style="padding:2px 8px 2px 0;">Std Dev (σ):</td><td style="font-weight:600;">{{series.options.sigma}}</td></tr>
                            <tr><td style="padding:2px 8px 2px 0;">Samples:</td><td style="font-weight:600;">{{series.options.count}}</td></tr>
                        </table>`
                }},
                plotOptions: {{
                    areaspline: {{
                        fillOpacity: 0.1,
                        lineWidth: 2,
                        stickyTracking: false,
                        marker: {{ enabled: false }},
                        states: {{ 
                            hover: {{ 
                                lineWidth: 4, 
                                fillOpacity: 0.3 
                            }},
                            inactive: {{
                                opacity: 0.3
                            }}
                        }}
                    }}
                }},
                series: []
            }});
        }}
        
        // Generate normal distribution curve dynamically
        function generateNormalCurve(mu, sigma, samplePoints) {{
            // Calculate bounds: μ ± 4σ covers ~99.99% of distribution
            const lowBound = mu - (4 * sigma);
            const highBound = mu + (4 * sigma);
            const step = (highBound - lowBound) / samplePoints;
            
            const points = [];
            for (let i = 0; i <= samplePoints; i++) {{
                const x = lowBound + (i * step);
                // Normal distribution PDF: (1 / (σ√(2π))) * e^(-0.5 * ((x-μ)/σ)²)
                const z = (x - mu) / sigma;
                const y = (1 / (sigma * Math.sqrt(2 * Math.PI))) * Math.exp(-0.5 * z * z);
                points.push([parseFloat(x.toFixed(4)), parseFloat(y.toFixed(6))]);
            }}
            
            return points;
        }}
        
        function plotSelectedCurves() {{
            const select = document.getElementById('circuit-select');
            const samplePoints = parseInt(document.getElementById('sample-points').value) || 50;
            
            // Validate sample points
            if (samplePoints < 10 || samplePoints > 300) {{
                alert('Sample points must be between 10 and 300');
                return;
            }}
            
            // Get selected circuit IDs
            const selectedIds = Array.from(select.selectedOptions).map(opt => opt.value);
            
            if (selectedIds.length === 0) {{
                alert('Please select at least one circuit');
                return;
            }}
            
            if (selectedIds.length > 50) {{
                alert('Maximum 50 circuits can be plotted at once for performance reasons');
                return;
            }}
            
            // Generate series data
            const seriesData = [];
            selectedIds.forEach(id => {{
                const metadata = curveMetadata.find(c => c.id === id);
                if (metadata) {{
                    const curvePoints = generateNormalCurve(metadata.mu, metadata.sigma, samplePoints);
                    seriesData.push({{
                        name: metadata.name,
                        data: curvePoints,
                        mu: metadata.mu,
                        sigma: metadata.sigma,
                        count: metadata.count,
                        turboThreshold: 0
                    }});
                }}
            }});
            
            // Update chart
            while(chart.series.length > 0) {{
                chart.series[0].remove(false);
            }}
            
            seriesData.forEach(series => {{
                chart.addSeries(series, false);
            }});
            
            chart.redraw();
            
            console.log(`Plotted ${{selectedIds.length}} curves with ${{samplePoints}} sample points each`);
        }}
        
        function selectAllCircuits() {{
            const select = document.getElementById('circuit-select');
            for (let option of select.options) {{
                option.selected = true;
            }}
        }}
        
        function clearSelection() {{
            const select = document.getElementById('circuit-select');
            for (let option of select.options) {{
                option.selected = false;
            }}
            
            // Clear chart
            while(chart.series.length > 0) {{
                chart.series[0].remove(false);
            }}
            chart.redraw();
        }}
        
        // Initialize on load
        window.onload = init;
        </script>
    </body>
    </html>
    """

    path = "circuit_bell_curves.html"
    with open(path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print("Interactive bell curve visualization saved")
    print(f"   Total circuits: {len(curve_metadata)}")
    print("   Sorted by: Mean (descending)")
    print("   Default selection: First 10 circuits (highest means)")

    return os.path.abspath(path)

if __name__ == "__main__":

    files = ["dataset/mlp_train.xlsx",
             "dataset/mlp_val.xlsx", "dataset/mlp_test.xlsx"]

    df = prep_raw_data("mlp/perfect_mlp.pth", "sae/sae_model.pth", files)

    save_styled_excel(df)

    # --- Usage Example ---
    matrix_df, metadata = generate_logic_heatmap_data(df)

    filepath = generate_logic_heatmap_html(matrix_df, metadata)
    print(f"Logic heatmap saved to: {filepath}")

    filepath = generate_stacked_norm_dist(df)
    print(f"Stacked norm dist saved to: {filepath}")

    filepath = generate_sankey_diagram_gchart(df)
    # print(f"Sankey diagram saved to: {filepath}")

    filepath = generate_sankey_diagram_plotly(df)
    # print(f"Sankey diagram saved to: {filepath}")

    # # --- Usage Example ---
    # matrix_df, metadata = generate_correlation_matrix(df)

    # filepath = generate_correlation_heatmap(matrix_df, metadata)
    # print(f"Correlation heatmap saved to: {filepath}")
